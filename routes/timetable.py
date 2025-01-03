# routes/timetable.py
# Стандартные библиотеки Python
import json
import os
import pickle
import tempfile
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Flask и связанные расширения
from flask import (Blueprint, Flask, render_template, jsonify, request, redirect, url_for, session, flash, send_file,
                   get_flashed_messages)
from flask_login import login_required
# Excel-related imports
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font)
from openpyxl.utils import get_column_letter
# Сторонние библиотеки
from transliterate import translit
from werkzeug.utils import secure_filename

# Локальные импорты
from config.config import Config
from models.history import TimetableHistory
from services.json_handler import TimetableHandler, SettingsHandler
from utils.decorators import admin_required
from utils.telegram_notifier import (notify_view, get_client_ip, send_lesson_change_notification)

bp = Blueprint('timetable', __name__, url_prefix='/timetable')

timetable_handler = TimetableHandler()
app = Flask(__name__)

app.config.from_object(Config)
settings_handler = SettingsHandler()

UPLOAD_FOLDER = 'data/uploads'
MERGED_FILE = 'data/timetable.json'


# --- РОУТЫ ---


@bp.route('/')
@notify_view
def index():
    """Главная страница"""
    timetable_data = timetable_handler.read_timetable()

    # Собираем информацию о неделях
    loaded_weeks = []
    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    loaded_weeks.append({'week_number': week.get('week_number'), 'date_start': week.get('date_start'),
                                         'date_end': week.get('date_end')})

    # Сортируем недели по номеру
    loaded_weeks.sort(key=lambda x: x['week_number'])

    # Собираем остальные данные
    groups = set()
    teachers = set()
    rooms = set()

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    for group in week.get('groups', []):
                        if 'group_name' in group:
                            groups.add(group['group_name'])

                        for day in group.get('days', []):
                            for lesson in day.get('lessons', []):
                                for teacher in lesson.get('teachers', []):
                                    if teacher.get('teacher_name'):
                                        teachers.add(teacher['teacher_name'])

                                for auditory in lesson.get('auditories', []):
                                    if auditory.get('auditory_name'):
                                        rooms.add(auditory['auditory_name'])

    return render_template('timetable/index.html', is_admin=session.get('is_admin', False), loaded_weeks=loaded_weeks,
                           flash_messages=get_flashed_messages(with_categories=True), groups=sorted(list(groups)),
                           teachers=sorted(list(teachers)), rooms=sorted(list(rooms)))


@bp.route('/overlaps')
@notify_view
def show_overlaps():
    """Страница отображения существующих накладок в расписании"""
    timetable_data = timetable_handler.read_timetable()
    all_overlaps = find_all_overlaps(timetable_data)
    settings = settings_handler.read_settings()

    return render_template('timetable/overlaps.html', overlaps=all_overlaps, settings=settings)


@bp.route('/api/settings/ignored_rooms', methods=['POST'])
@login_required
@admin_required
def manage_ignored_rooms():
    data = request.get_json()
    settings = settings_handler.read_settings()

    action = data.get('action')
    room = data.get('room')

    if action == 'add' and room:
        if room not in settings['ignored_rooms']:
            settings['ignored_rooms'].append(room)
    elif action == 'remove' and room:
        if room in settings['ignored_rooms']:
            settings['ignored_rooms'].remove(room)

    if settings_handler.save_settings(settings):
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Failed to save settings'}), 500


@bp.route('/api/transfer_options', methods=['POST'])
def get_transfer_options():
    try:
        data = request.get_json()

        group_name = data.get('group_name')
        current_day = data.get('current_day')
        current_time = data.get('current_time')
        current_week = data.get('current_week')
        lesson_data = data.get('lesson')

        if not all([group_name, current_day, current_time, current_week, lesson_data]):
            return jsonify({'error': 'Отсутствуют необходимые параметры'}), 400

        timetable_data = timetable_handler.read_timetable()

        options = find_transfer_slots(timetable_data, group_name, lesson_data, current_week, current_day, current_time)

        return jsonify(options)

    except Exception as e:
        print(f"Ошибка в get_transfer_options: {str(e)}")
        return jsonify({'error': f'Внутренняя ошибка сервера: {str(e)}'}), 500


@bp.route('/export/<type>/<name>')
def export_excel(type, name):
    """Экспорт расписания в Excel"""
    try:
        week = request.args.get('week')
        if not week:
            return "Не указана неделя для экспорта", 400

        timetable_data = timetable_handler.read_timetable()
        exporter = ExcelExporter()

        if type == 'group':
            wb = exporter.create_excel(timetable_data, week, group_name=name)
            filename = f"Расписание_группы_{name}_неделя_{week}.xlsx"
        elif type == 'teacher':
            wb = exporter.create_excel(timetable_data, week, teacher_name=name)
            filename = f"Расписание_преподавателя_{name}_неделя_{week}.xlsx"
        elif type == 'room':
            wb = exporter.create_excel(timetable_data, week, room_name=name)
            filename = f"Расписание_аудитории_{name}_неделя_{week}.xlsx"
        else:
            return "Неверный тип экспорта", 400

        # Конвертируем имя файла в безопасный формат и транслитерируем русские буквы
        filename = translit(filename, 'ru', reversed=True)
        filename = secure_filename(filename)

        # Сохраняем в BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"Ошибка при экспорте в Excel: {str(e)}")
        return f"Ошибка при создании Excel файла: {str(e)}", 500


@bp.route('/group/<group_name>')
@notify_view
def group_timetable(group_name):
    timetable_data = timetable_handler.read_timetable()
    selected_week = request.args.get('week', None)

    # Получаем список всех недель
    weeks = []
    if isinstance(timetable_data, list):
        for timetable_item in timetable_data:
            if 'timetable' in timetable_item:
                for week_data in timetable_item['timetable']:
                    weeks.append(
                        {'week_number': week_data.get('week_number'), 'date_start': week_data.get('date_start'),
                         'date_end': week_data.get('date_end')})

    # Сортируем недели по номеру
    weeks.sort(key=lambda x: x['week_number'])

    # Если неделя не выбрана, определяем текущую
    if not selected_week and weeks:
        current_date = datetime.now()
        current_week_num = current_date.isocalendar()[1]

        closest_week = weeks[0]
        min_diff = abs(datetime.strptime(weeks[0]['date_start'], '%d-%m-%Y').isocalendar()[1] - current_week_num)

        for week in weeks:
            try:
                week_start = datetime.strptime(week['date_start'], '%d-%m-%Y')
                week_diff = abs(week_start.isocalendar()[1] - current_week_num)
                if week_diff < min_diff:
                    min_diff = week_diff
                    closest_week = week
            except ValueError:
                continue

        selected_week = str(closest_week['week_number'])
        return redirect(url_for('timetable.group_timetable', group_name=group_name, week=selected_week))

    # Расписание звонков
    time_slots = [{'start': '08:00', 'end': '09:20'}, {'start': '09:30', 'end': '10:50'},
                  {'start': '11:00', 'end': '12:20'}, {'start': '12:40', 'end': '14:00'},
                  {'start': '14:10', 'end': '15:30'}, {'start': '15:40', 'end': '17:00'},
                  {'start': '17:10', 'end': '18:30'}, {'start': '18:40', 'end': '20:00'}]

    day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

    # Получаем текущую дату и время
    current_datetime = datetime.now()
    current_time = current_datetime.strftime('%H:%M')
    current_weekday = current_datetime.weekday() + 1

    # Определяем текущую пару
    current_pair = None
    for i, time_slot in enumerate(time_slots, 1):
        start_time = datetime.strptime(time_slot['start'], '%H:%M').time()
        end_time = datetime.strptime(time_slot['end'], '%H:%M').time()
        if start_time <= current_datetime.time() <= end_time:
            current_pair = i
            break

    # Получаем даты для дней недели
    dates = [''] * 6  # По умолчанию пустые строки для дат
    if selected_week and weeks:
        week_data = next((week for week in weeks if str(week['week_number']) == str(selected_week)), None)
        if week_data and week_data.get('date_start'):
            try:
                date_formats = ['%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d']
                start_date = None

                for date_format in date_formats:
                    try:
                        start_date = datetime.strptime(week_data['date_start'], date_format)
                        break
                    except ValueError:
                        continue

                if start_date:
                    dates = []
                    for i in range(6):  # 6 дней недели
                        date = start_date + timedelta(days=i)
                        dates.append(date.strftime('%d.%m'))
            except Exception as e:
                print(f"Ошибка при парсинге даты: {week_data['date_start']}, ошибка: {str(e)}")
                dates = [''] * 6

    # Получаем уникальные значения для выпадающих списков
    unique_values = get_unique_values(timetable_data)

    # Создаем функцию-обертку для передачи в шаблон
    def get_lessons_wrapper(timetable, day, time):
        return get_lessons(timetable, day, time, group_name, selected_week)

    return render_template('timetable/group.html', group_name=group_name, weeks=weeks, current_week=selected_week,
                           time_slots=time_slots, day_names=day_names, dates=dates, current_weekday=current_weekday,
                           current_pair=current_pair, current_time=current_time, timetable=timetable_data,
                           get_lessons=get_lessons_wrapper, unique_values=unique_values)


@bp.route('/free_rooms', methods=['GET', 'POST'])
@notify_view
def free_rooms():
    """Поиск свободных аудиторий"""
    timetable_data = timetable_handler.read_timetable()

    # Get all rooms and extract buildings
    all_rooms = set()
    buildings = {'all': 'Все корпуса', 'other': 'Остальное'}

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    for group in week.get('groups', []):
                        for day in group.get('days', []):
                            for lesson in day.get('lessons', []):
                                for auditory in lesson.get('auditories', []):
                                    room = auditory.get('auditory_name')
                                    if room:
                                        all_rooms.add(room)
                                        # Extract building number
                                        if '.' in room:
                                            building = room.split('.')[0]
                                            if building.isdigit():
                                                buildings[building] = f'Корпус {building}'

    weeks = []
    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    weeks.append({'week_number': week.get('week_number'), 'date_start': week.get('date_start'),
                                  'date_end': week.get('date_end')})
    weeks.sort(key=lambda x: x['week_number'])

    free_rooms = []
    selected_week = None
    selected_day = None
    selected_time = None
    selected_building = None

    if request.method == 'POST':
        selected_week = request.form.get('week')
        selected_day = request.form.get('day')
        selected_time = request.form.get('time')
        selected_building = request.form.get('building', 'all')

        if all([selected_week, selected_day, selected_time]):
            busy_rooms = set()
            for data in timetable_data:
                if 'timetable' in data:
                    for week in data['timetable']:
                        if str(week.get('week_number')) == selected_week:
                            for group in week.get('groups', []):
                                for day in group.get('days', []):
                                    if str(day.get('weekday')) == selected_day:
                                        for lesson in day.get('lessons', []):
                                            if str(lesson.get('time')) == selected_time:
                                                for auditory in lesson.get('auditories', []):
                                                    if auditory.get('auditory_name'):
                                                        busy_rooms.add(auditory['auditory_name'])

            # Filter rooms by building
            available_rooms = all_rooms - busy_rooms
            if selected_building != 'all':
                if selected_building == 'other':
                    free_rooms = [room for room in available_rooms if not room.split('.')[0].isdigit()]
                else:
                    free_rooms = [room for room in available_rooms if room.startswith(f"{selected_building}.")]
            else:
                free_rooms = sorted(list(available_rooms))

    time_slots = [{'start': '08:00', 'end': '09:20'}, {'start': '09:30', 'end': '10:50'},
                  {'start': '11:00', 'end': '12:20'}, {'start': '12:40', 'end': '14:00'},
                  {'start': '14:10', 'end': '15:30'}, {'start': '15:40', 'end': '17:00'},
                  {'start': '17:10', 'end': '18:30'}, {'start': '18:40', 'end': '20:00'}]

    day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

    return render_template('timetable/free_rooms.html', weeks=weeks, time_slots=time_slots, day_names=day_names,
                           free_rooms=free_rooms, selected_week=selected_week, selected_day=selected_day,
                           selected_time=selected_time, selected_building=selected_building,
                           buildings=dict(sorted(buildings.items())), all_rooms=sorted(list(all_rooms)))


@bp.route('/api/group/<group_name>')
def get_group_timetable_json(group_name):
    """API для получения расписания группы в формате JSON"""
    timetable = timetable_handler.get_group_timetable(group_name)
    if timetable:
        return jsonify(timetable)
    return jsonify({"error": "Group not found"}), 404


@bp.route('/api/update', methods=['POST'])
@login_required
@admin_required
def update_timetable():
    """API для обновления расписания с проверкой накладок"""
    try:
        data = request.get_json()
        print("\n=== DEBUG: RECEIVED DATA ===")
        print(json.dumps(data, indent=2, ensure_ascii=False))

        history_handler = TimetableHistory()

        group_name = data.get('group_name')
        day = data.get('day')
        time = data.get('time')
        lessons = data.get('lessons', [])
        week = request.args.get('week')
        ignore_overlaps = data.get('ignore_overlaps', False)

        if not week:
            return jsonify({"status": "error", "error": "Не указан номер недели"}), 400

        print("Данные успешно загружены")

        # Получаем текущие данные напрямую из базы
        current_lessons = []
        timetable_data = timetable_handler.read_timetable()

        # Если это не игнорирование накладок, проверяем их
        if not ignore_overlaps:
            overlaps = check_overlaps(timetable_data, week, day, time, lessons, group_name)
            if overlaps['total_count'] > 0:
                warning_message = "Обнаружены следующие накладки:\n"

                for room_overlap in overlaps['room_overlaps']:
                    warning_message += f"\nНакладка в аудитории {room_overlap['room']}:\n"
                    for lesson in room_overlap['lessons']:
                        warning_message += f"- {lesson['subject']} ({lesson['group']}, преп. {lesson['teacher']})\n"

                for teacher_overlap in overlaps['teacher_overlaps']:
                    warning_message += f"\nНакладка у преподавателя {teacher_overlap['teacher']}:\n"
                    for lesson in teacher_overlap['lessons']:
                        warning_message += f"- {lesson['subject']} ({lesson['group']}, ауд. {lesson['room']})\n"

                return jsonify({"status": "warning", "message": warning_message, "overlaps": overlaps}), 409

        # Получаем текущие занятия для записи в историю
        if isinstance(timetable_data, list):
            for item in timetable_data:
                if 'timetable' in item:
                    for week_data in item['timetable']:
                        if str(week_data.get('week_number')) == str(week):
                            for group in week_data.get('groups', []):
                                if group.get('group_name') == group_name:
                                    for day_data in group.get('days', []):
                                        if day_data.get('weekday') == day:
                                            current_lessons.extend([lesson for lesson in day_data.get('lessons', []) if
                                                                    lesson.get('time') == time])

        print("\n=== DEBUG: CURRENT STATE ===")
        print("Current lessons in database:")
        print(json.dumps(current_lessons, indent=2, ensure_ascii=False))
        print("\nNew lessons from request:")
        print(json.dumps(lessons, indent=2, ensure_ascii=False))

        # Определяем тип операции
        if not current_lessons and lessons:
            action = 'create'
            print("\nACTION: CREATE - Adding new lessons")
        elif current_lessons and not lessons:
            action = 'delete'
            print("\nACTION: DELETE - Removing all lessons")
        else:
            # Проверяем, действительно ли изменились данные
            changes = False
            if len(lessons) != len(current_lessons):
                changes = True
            else:
                for new, old in zip(lessons, current_lessons):
                    if (new.get('subject') != old.get('subject') or new.get('type') != old.get('type') or new.get(
                            'subgroup') != old.get('subgroup') or str(new.get('teachers')) != str(
                        old.get('teachers')) or str(new.get('auditories')) != str(old.get('auditories'))):
                        changes = True
                        break

            action = 'update' if changes else None
            print(f"\nACTION: {action or 'NO CHANGE NEEDED'}")

        print(f"\nFinal determined action: {action}")

        if action:
            # Добавляем номер недели к каждому новому занятию
            for lesson in lessons:
                lesson['week'] = int(week)

            # Подготавливаем данные для истории
            history_data = {'group_name': group_name, 'day': day, 'time': time, 'week': week, 'lessons': lessons,
                            'old_lessons': current_lessons, 'editor_ip': get_client_ip()}

            # Добавляем запись в историю
            history_handler.add_record(action, history_data)

            # Обновляем данные в базе
            updated = False
            if isinstance(timetable_data, list):
                for item in timetable_data:
                    if 'timetable' in item:
                        for week_data in item['timetable']:
                            if str(week_data.get('week_number')) == str(week):
                                print(f"Обновление занятий для группы {group_name}, день {day}, время {time}")
                                print(f"Новые занятия: {lessons}")
                                for group in week_data.get('groups', []):
                                    if group.get('group_name') == group_name:
                                        print(f"Найдена группа {group_name}")
                                        for day_data in group.get('days', []):
                                            if day_data.get('weekday') == day:
                                                print(f"Найден день {day}")
                                                # Сохраняем существующие занятия для других пар
                                                other_lessons = [lesson for lesson in day_data.get('lessons', []) if
                                                                 lesson.get('time') != time]

                                                # Сохраняем занятия для других недель
                                                other_week_lessons = [lesson for lesson in day_data.get('lessons', [])
                                                                      if
                                                                      lesson.get('time') == time and lesson.get('week',
                                                                                                                int(week)) != int(
                                                                          week)]

                                                # Добавляем новые занятия
                                                day_data['lessons'] = other_lessons + other_week_lessons + lessons
                                                print(f"Занятия обновлены")
                                                updated = True
                                                break
                        if updated:
                            break

            if not updated:
                return jsonify({"status": "error", "error": "Failed to update timetable"}), 500

            # Создаем резервную копию перед сохранением
            timetable_handler._create_backup()
            print("Создана резервная копия")

            # Сохраняем обновленные данные
            success = timetable_handler.save_timetable(timetable_data)

            if success:
                # Отправляем уведомления
                if action == 'create':
                    for lesson in lessons:
                        send_lesson_change_notification(action='create', group_name=group_name, weekday=day,
                                                        time_slot=time, week_number=week, lesson_data=lesson,
                                                        editor_ip=get_client_ip())
                elif action == 'delete':
                    for lesson in current_lessons:
                        send_lesson_change_notification(action='delete', group_name=group_name, weekday=day,
                                                        time_slot=time, week_number=week, lesson_data=lesson,
                                                        editor_ip=get_client_ip())
                elif action == 'update':
                    for i, new_lesson in enumerate(lessons):
                        if i < len(current_lessons):
                            send_lesson_change_notification(action='update', group_name=group_name, weekday=day,
                                                            time_slot=time, week_number=week, lesson_data=new_lesson,
                                                            old_lesson_data=current_lessons[i],
                                                            editor_ip=get_client_ip())

                return jsonify({"status": "success"})
            else:
                return jsonify({"status": "error", "error": "Failed to save timetable"}), 500
        else:
            return jsonify({"status": "success", "message": "No changes needed"})

    except Exception as e:
        print(f"Error updating timetable: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return jsonify({"status": "error", "error": str(e)}), 500


@bp.route('/search', methods=['GET', 'POST'])
@notify_view
def search_timetable():
    """Поиск занятий по параметрам"""
    # Инициализация начальных значений
    groups = set()
    subjects = set()
    lesson_types = {'л.', 'пр.', 'лаб.'}
    search_results = []
    date_range = None

    # Обновленные параметры поиска
    search_params = {'group': request.form.get('group', ''), 'subject': request.form.get('subject', ''),
                     'lesson_type': request.form.get('lesson_type', '')} if request.method == 'POST' else {'group': '',
                                                                                                           'subject': '',
                                                                                                           'lesson_type': ''}

    if request.method == 'POST':
        # Загружаем данные только при POST-запросе
        timetable_handler = TimetableHandler()
        timetable_data = timetable_handler.read_timetable()
        all_dates = []

        # Получаем параметры поиска
        export_format = request.form.get('export_format')

        # Собираем уникальные значения и выполняем поиск
        if isinstance(timetable_data, list):
            for data in timetable_data:
                if isinstance(data, dict) and 'timetable' in data:
                    for week in data['timetable']:
                        # Сбор дат для диапазона
                        try:
                            start_date = parse_date(week.get('date_start', ''))
                            end_date = parse_date(week.get('date_end', ''))
                            if start_date and end_date:
                                all_dates.extend([start_date, end_date])
                        except Exception as e:
                            print(f"Ошибка обработки дат: {e}")
                            continue

                        week_number = week.get('week_number')
                        date_start = week.get('date_start')
                        date_end = week.get('date_end')

                        # Обработка групп и поиск
                        for group_data in week.get('groups', []):
                            group_name = group_data.get('group_name')
                            if group_name:
                                groups.add(group_name)

                            # Пропускаем, если группа не соответствует
                            if search_params['group'] and group_name != search_params['group']:
                                continue

                            for day in group_data.get('days', []):
                                weekday = day.get('weekday')
                                if weekday:
                                    for lesson in day.get('lessons', []):
                                        # Сбор предметов
                                        subject = lesson.get('subject')
                                        if subject:
                                            subjects.add(subject)

                                        # Проверка соответствия критериям поиска
                                        if (search_params['subject'] and lesson.get('subject') != search_params[
                                            'subject']) or (
                                                search_params['lesson_type'] and lesson.get('type') != search_params[
                                            'lesson_type']):
                                            continue

                                        # Получаем дату занятия
                                        start_date = parse_date(date_start)
                                        if not start_date:
                                            continue

                                        lesson_date = start_date + timedelta(days=weekday - 1)

                                        # Добавляем результат с подгруппой
                                        search_results.append({'date': lesson_date.strftime('%d.%m.%Y'), 'weekday':
                                            ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота'][
                                                weekday - 1], 'week': week_number,
                                                               'week_period': f"{date_start} - {date_end}",
                                                               'time': f"Пара {lesson.get('time')}",
                                                               'time_number': lesson.get('time'), 'group': group_name,
                                                               'subject': lesson.get('subject'),
                                                               'type': lesson.get('type'),
                                                               'subgroup': lesson.get('subgroup', 0),  # Добавлено
                                                               'teachers': [t.get('teacher_name') for t in
                                                                            lesson.get('teachers', [])],
                                                               'auditories': [a.get('auditory_name') for a in
                                                                              lesson.get('auditories', [])]})

        # Определяем диапазон дат
        if all_dates:
            min_date = min(all_dates).strftime('%d.%m.%Y')
            max_date = max(all_dates).strftime('%d.%m.%Y')
            date_range = f"с {min_date} по {max_date}"

        # Сортируем результаты
        search_results.sort(key=lambda x: (
            datetime.strptime(x['date'], '%d.%m.%Y'), x['time_number'], x['group'],  # Добавлена сортировка по группе
            x['subgroup'] or 0  # Добавлена сортировка по подгруппе
        ))

        if export_format == 'excel' and search_results:
            return export_search_results(search_results)

    # При GET-запросе загружаем только списки для выпадающих меню
    if request.method == 'GET':
        timetable_handler = TimetableHandler()
        timetable_data = timetable_handler.read_timetable()

        if isinstance(timetable_data, list):
            for data in timetable_data:
                if 'timetable' in data:
                    for week in data['timetable']:
                        for group in week.get('groups', []):
                            if group.get('group_name'):
                                groups.add(group.get('group_name'))
                            for day in group.get('days', []):
                                for lesson in day.get('lessons', []):
                                    if lesson.get('subject'):
                                        subjects.add(lesson.get('subject'))

    return render_template('timetable/search.html', groups=sorted(list(filter(None, groups))),
                           subjects=sorted(list(filter(None, subjects))), lesson_types=sorted(list(lesson_types)),
                           search_results=search_results, date_range=date_range, search_params=search_params)


def parse_date(date_str):
    """
    Преобразует строку даты в объект datetime с поддержкой различных форматов.

    Пытается разобрать дату в нескольких распространенных форматах.

    Args:
        date_str (str): Строка с датой в одном из поддерживаемых форматов

    Returns:
        datetime или None: Объект datetime если парсинг успешен, None если дата не распознана

    Supported formats:
        - DD-MM-YYYY
        - DD.MM.YYYY
        - YYYY-MM-DD
    """

    formats = ['%d-%m-%Y', '%d.%m.%Y', '%Y-%m-%d']
    for date_format in formats:
        try:
            return datetime.strptime(date_str, date_format)
        except ValueError:
            continue
    return None


def export_search_results(results):
    """
    Экспортирует результаты поиска в Excel файл.

    Создает форматированную Excel таблицу с результатами поиска по расписанию,
    включая стилизацию и автоматическую настройку размеров.

    Args:
        results (list): Список результатов поиска, каждый элемент содержит:
            - date: Дата занятия
            - weekday: День недели
            - week: Номер недели
            - week_period: Период недели
            - time: Время занятия
            - group: Группа
            - subject: Предмет
            - type: Тип занятия
            - teachers: Список преподавателей
            - auditories: Список аудиторий

    Returns:
        flask.Response: Ответ с Excel файлом для скачивания

    Note:
        Создает временный файл в памяти
        Применяет форматирование: шрифты, границы, выравнивание
        Автоматически настраивает ширину столбцов
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты поиска"

    # Заголовки
    headers = ['Дата', 'День недели', 'Номер недели', 'Период недели', 'Время', 'Группа', 'Дисциплина', 'Тип занятия',
               'Подгруппа', 'Преподаватели', 'Аудитории']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

    # Данные
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['date'])
        ws.cell(row=row, column=2, value=result['weekday'])
        ws.cell(row=row, column=3, value=f"Неделя {result['week']}")
        ws.cell(row=row, column=4, value=result['week_period'])
        ws.cell(row=row, column=5, value=result['time'])
        ws.cell(row=row, column=6, value=result['group'])
        ws.cell(row=row, column=7, value=result['subject'])
        ws.cell(row=row, column=8, value=result['type'])
        ws.cell(row=row, column=9, value=f"Подгруппа {result['subgroup']}" if result['subgroup'] > 0 else "Общее")
        ws.cell(row=row, column=10, value=', '.join(result['teachers']))
        ws.cell(row=row, column=11, value=', '.join(result['auditories']))

    # Настройка ширины столбцов
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Применяем форматирование к данным
    for row in range(2, len(results) + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    # Создаем временный файл
    with BytesIO() as excel_file:
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(excel_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'search_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')


@bp.route('/teacher/<teacher_name>')
@notify_view
def teacher_timetable(teacher_name):
    timetable_data = timetable_handler.read_timetable()
    selected_week = request.args.get('week', None)

    weeks = []
    if isinstance(timetable_data, list):
        for timetable_item in timetable_data:
            if 'timetable' in timetable_item:
                for week_data in timetable_item['timetable']:
                    weeks.append(
                        {'week_number': week_data.get('week_number'), 'date_start': week_data.get('date_start'),
                         'date_end': week_data.get('date_end')})

    weeks.sort(key=lambda x: x['week_number'])

    if not selected_week and weeks:
        current_date = datetime.now()
        current_week_num = current_date.isocalendar()[1]

        closest_week = weeks[0]
        min_diff = float('inf')

        for week in weeks:
            try:
                try:
                    week_start = datetime.strptime(week['date_start'], '%d.%m.%Y')
                except ValueError:
                    week_start = datetime.strptime(week['date_start'], '%d-%m-%Y')

                week_diff = abs(week_start.isocalendar()[1] - current_week_num)
                if week_diff < min_diff:
                    min_diff = week_diff
                    closest_week = week
            except ValueError:
                continue

        selected_week = str(closest_week['week_number'])
        return redirect(url_for('timetable.teacher_timetable', teacher_name=teacher_name, week=selected_week))

    time_slots = [{'start': '08:00', 'end': '09:20'}, {'start': '09:30', 'end': '10:50'},
                  {'start': '11:00', 'end': '12:20'}, {'start': '12:40', 'end': '14:00'},
                  {'start': '14:10', 'end': '15:30'}, {'start': '15:40', 'end': '17:00'},
                  {'start': '17:10', 'end': '18:30'}, {'start': '18:40', 'end': '20:00'}]

    day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

    current_datetime = datetime.now()
    current_time = current_datetime.strftime('%H:%M')
    current_weekday = current_datetime.weekday() + 1

    current_pair = None
    for i, time_slot in enumerate(time_slots, 1):
        start_time = datetime.strptime(time_slot['start'], '%H:%M').time()
        end_time = datetime.strptime(time_slot['end'], '%H:%M').time()
        if start_time <= current_datetime.time() <= end_time:
            current_pair = i
            break

    def get_teacher_lessons(timetable, day, time):
        lessons_by_type = {}

        try:
            if isinstance(timetable, list):
                for data in timetable:
                    if 'timetable' in data:
                        for week in data['timetable']:
                            if selected_week and str(week.get('week_number')) != str(selected_week):
                                continue

                            for group in week.get('groups', []):
                                group_name = group.get('group_name')
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            if (lesson.get('time') == time and any(
                                                    teacher.get('teacher_name') == teacher_name for teacher in
                                                    lesson.get('teachers', []))):

                                                key = (lesson.get('subject'), lesson.get('type'),
                                                       lesson.get('auditories', [{}])[0].get('auditory_name'))

                                                if key not in lessons_by_type:
                                                    lesson_copy = lesson.copy()
                                                    lesson_copy['groups'] = [group_name]
                                                    lessons_by_type[key] = lesson_copy
                                                else:
                                                    lessons_by_type[key]['groups'].append(group_name)

            result = list(lessons_by_type.values())
            for lesson in result:
                lesson['groups'].sort()

            return result if result else None

        except Exception as e:
            print(f"Ошибка в get_teacher_lessons: {str(e)}")
            return None

    return render_template('timetable/teacher.html', teacher_name=teacher_name, weeks=weeks, current_week=selected_week,
                           time_slots=time_slots, day_names=day_names, timetable=timetable_data,
                           get_lessons=get_teacher_lessons, current_pair=current_pair, current_weekday=current_weekday,
                           current_time=current_time)


@bp.route('/api/subjects_by_group/<group_name>')
def get_subjects_by_group(group_name):
    """Получение списка предметов для конкретной группы"""
    timetable_data = timetable_handler.read_timetable()
    subjects = set()

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    for group in week.get('groups', []):
                        if group.get('group_name') == group_name:
                            for day in group.get('days', []):
                                for lesson in day.get('lessons', []):
                                    if lesson.get('subject'):
                                        subjects.add(lesson.get('subject'))

    return jsonify(sorted(list(subjects)))


@bp.route('/login', methods=['POST'])
def login():
    """Простая авторизация администратора"""
    password = request.form.get('password')
    if password == app.config['ADMIN_PASSWORD']:
        session['is_admin'] = True
        flash('Вы успешно вошли как администратор', 'success')
    else:
        flash('Неверный пароль', 'error')
    return redirect(url_for('timetable.index'))


@bp.route('/logout')
def logout():
    """Выход из админки"""
    session.pop('is_admin', None)
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('timetable.index'))


@bp.route('/upload', methods=['POST'])
@admin_required
def upload_files():
    """Загрузка файлов расписания"""
    global content
    print("\n=== Начало загрузки файлов ===")

    if 'timetable_files' not in request.files:
        print("Файлы не найдены в запросе")
        flash('Нет выбранных файлов', 'error')
        return redirect(url_for('timetable.index'))

    files = request.files.getlist('timetable_files')
    if not files:
        print("Список файлов пуст")
        flash('Нет выбранных файлов', 'error')
        return redirect(url_for('timetable.index'))

    print(f"Получено файлов: {len(files)}")
    print(f"Имена файлов: {[f.filename for f in files]}")

    # Структуры для хранения данных
    file_data_list = []  # Данные из всех файлов
    existing_weeks = set()  # Существующие номера недель
    conflicts = []  # Информация о конфликтах

    # Читаем существующие данные
    current_data = read_merged_file()
    print(f"Текущие данные загружены: {current_data is not None}")

    # Собираем информацию о существующих неделях
    if current_data:
        if isinstance(current_data, list):
            for item in current_data:
                if 'timetable' in item:
                    for week in item['timetable']:
                        existing_weeks.add(week['week_number'])
                    file_data_list.extend(current_data)
        print(f"Найдено существующих недель: {len(existing_weeks)}")
        print(f"Номера недель: {sorted(list(existing_weeks))}")

    # Обрабатываем новые файлы
    for file in files:
        if not file.filename:
            continue

        print(f"\nОбработка файла: {file.filename}")
        try:
            # Читаем содержимое файла
            file_content = file.read()
            print(f"Размер файла: {len(file_content)} байт")

            # Пробуем разные кодировки
            decoded = False
            for encoding in ['windows-1251', 'utf-8', 'utf-8-sig']:
                try:
                    content = file_content.decode(encoding)
                    print(f"Успешная декодировка с {encoding}")
                    decoded = True
                    break
                except UnicodeDecodeError:
                    continue

            if not decoded:
                raise UnicodeDecodeError("Не удалось определить кодировку файла")

            # Обрабатываем escape-последовательности
            content = content.replace('\\', '\\\\')

            # Парсим JSON
            file_data = json.loads(content)
            print("JSON успешно прочитан")

            if not isinstance(file_data, list):
                file_data = [file_data]

            # Проверяем недели в файле
            for item in file_data:
                if 'timetable' in item:
                    for week in item['timetable']:
                        week_num = week['week_number']
                        if week_num in existing_weeks:
                            print(f"Конфликт: неделя {week_num} в файле {file.filename}")
                            conflicts.append({'week': week_num, 'file': file.filename, 'date_start': week['date_start'],
                                              'date_end': week['date_end']})
                file_data_list.append(item)

        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга JSON: {str(e)}")
            if hasattr(e, 'pos'):
                start = max(0, e.pos - 50)
                end = min(len(content), e.pos + 50)
                print(f"Контекст ошибки: {content[start:end]}")
            flash(f'Ошибка парсинга JSON в файле {file.filename}', 'error')
            return redirect(url_for('timetable.index'))
        except Exception as e:
            print(f"Ошибка обработки файла: {str(e)}")
            flash(f'Ошибка при обработке файла {file.filename}: {str(e)}', 'error')
            return redirect(url_for('timetable.index'))

    print(f"\nИтоги обработки:")
    print(f"Всего собрано данных: {len(file_data_list)}")
    print(f"Конфликтов найдено: {len(conflicts)}")

    # Если есть конфликты
    if conflicts:
        print("Сохранение данных во временный файл")
        temp_id = save_temp_data({'pending_data': file_data_list, 'conflicts': conflicts})
        session['temp_id'] = temp_id
        return redirect(url_for('timetable.resolve_conflicts'))

    try:
        print("Объединение и сохранение данных")
        # Собираем все недели
        all_weeks = []
        for item in file_data_list:
            if 'timetable' in item:
                all_weeks.extend(item['timetable'])

        # Объединяем недели
        merged_weeks = merge_weeks(all_weeks)
        final_data = [{'timetable': merged_weeks}]

        # Сохраняем результат
        if save_merged_data(final_data):
            flash('Файлы успешно загружены и объединены', 'success')
        else:
            flash('Ошибка при сохранении данных', 'error')

    except Exception as e:
        print(f"Ошибка при объединении данных: {str(e)}")
        flash(f'Ошибка при объединении данных: {str(e)}', 'error')

    print("=== Завершение загрузки файлов ===\n")
    return redirect(url_for('timetable.index'))


@bp.route('/teacher/<teacher_name>/workload/export')
@notify_view
def export_teacher_workload_route(teacher_name):
    timetable_data = timetable_handler.read_timetable()
    workload_data = {}

    def get_week_dates(group_data):
        """Вычисляет даты недели для группы на основе её расписания"""
        week_dates = {}
        min_day = 7  # для поиска первого дня недели
        max_day = 0  # для поиска последнего дня недели

        # Ищем крайние дни недели в расписании группы
        for day in group_data.get('days', []):
            day_num = day.get('weekday', 0)
            if day.get('lessons'):  # Проверяем, есть ли занятия в этот день
                min_day = min(min_day, day_num)
                max_day = max(max_day, day_num)

        return min_day, max_day

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    week_num = week.get('week_number')
                    base_start_date = week.get('date_start', '')

                    if base_start_date and week_num:
                        try:
                            # Преобразуем базовую дату начала недели
                            if '-' in base_start_date:
                                base_date = datetime.strptime(base_start_date, '%d-%m-%Y')
                            else:
                                base_date = datetime.strptime(base_start_date, '%d.%m.%Y')

                            for group in week.get('groups', []):
                                group_name = group.get('group_name')

                                # Получаем крайние дни недели для группы
                                min_day, max_day = get_week_dates(group)

                                if min_day < 7 and max_day > 0:  # Если нашли дни с занятиями
                                    # Вычисляем даты для этой группы
                                    group_start_date = base_date + timedelta(days=min_day - 1)
                                    group_end_date = base_date + timedelta(days=max_day - 1)
                                    week_dates = f"{group_start_date.strftime('%d.%m.%Y')} {group_end_date.strftime('%d.%m.%Y')}"
                                else:
                                    week_dates = ''

                                for day in group.get('days', []):
                                    for lesson in day.get('lessons', []):
                                        if any(teacher.get('teacher_name') == teacher_name for teacher in
                                               lesson.get('teachers', [])):
                                            subject = lesson.get('subject')

                                            if subject not in workload_data:
                                                workload_data[subject] = {'groups': {}}

                                            if group_name not in workload_data[subject]['groups']:
                                                workload_data[subject]['groups'][group_name] = {'schedule': []}

                                            lesson_info = {
                                                'week': week_num,
                                                'type': lesson.get('type'),
                                                'week_dates': week_dates
                                            }
                                            workload_data[subject]['groups'][group_name]['schedule'].append(lesson_info)

                            # После обработки всех занятий добавляем информацию о неделях без занятий
                            for subject_data in workload_data.values():
                                for group_name, group_data in subject_data['groups'].items():
                                    # Ищем группу в текущей неделе
                                    current_group = next((g for g in week.get('groups', [])
                                                          if g.get('group_name') == group_name), None)

                                    if current_group:
                                        min_day, max_day = get_week_dates(current_group)
                                        if min_day < 7 and max_day > 0:
                                            group_start_date = base_date + timedelta(days=min_day - 1)
                                            group_end_date = base_date + timedelta(days=max_day - 1)
                                            week_dates = f"{group_start_date.strftime('%d.%m.%Y')} {group_end_date.strftime('%d.%m.%Y')}"
                                        else:
                                            week_dates = ''

                                        existing_weeks = {lesson['week'] for lesson in group_data['schedule']}
                                        if week_num not in existing_weeks:
                                            group_data['schedule'].append({
                                                'week': week_num,
                                                'type': None,
                                                'week_dates': week_dates
                                            })
                        except Exception as e:
                            print(f"Error processing dates: {str(e)}")
                            continue

    return export_teacher_workload(teacher_name, workload_data)


@bp.route('/resolve_conflicts')
@admin_required
def resolve_conflicts():
    """Страница разрешения конфликтов"""
    temp_id = session.get('temp_id')
    if not temp_id:
        flash('Нет конфликтов для разрешения', 'info')
        return redirect(url_for('timetable.index'))

    temp_data = load_temp_data(temp_id)
    if not temp_data or 'conflicts' not in temp_data:
        flash('Данные о конфликтах не найдены', 'error')
        return redirect(url_for('timetable.index'))

    # Группируем конфликты по номерам недель
    conflicts_by_week = {}
    for conflict in temp_data['conflicts']:
        week_num = conflict['week']
        if week_num not in conflicts_by_week:
            conflicts_by_week[week_num] = {'week': week_num, 'date_start': conflict['date_start'],
                                           'date_end': conflict['date_end'], 'files': []}
        conflicts_by_week[week_num]['files'].append(conflict['file'])

    return render_template('timetable/resolve_conflicts.html', conflicts=conflicts_by_week.values())


@bp.route('/apply_resolution', methods=['POST'])
@admin_required
def apply_resolution():
    """Применение решений по конфликтам"""
    resolutions = request.json
    temp_id = session.get('temp_id')

    if not temp_id:
        return jsonify({'status': 'error', 'message': 'Данные для обработки не найдены'}), 400

    temp_data = load_temp_data(temp_id)
    if not temp_data:
        return jsonify({'status': 'error', 'message': 'Временные данные не найдены'}), 400

    pending_data = temp_data['pending_data']
    current_data = read_merged_file() or []

    # Обрабатываем каждую неделю согласно решению
    for week_num, action in resolutions.items():
        week_num = int(week_num)

        if action == 'skip':
            # Пропускаем новые данные для этой недели
            pending_data = [item for item in pending_data if
                            'timetable' not in item or not any(w['week_number'] == week_num for w in item['timetable'])]
        elif action == 'replace':
            # Удаляем старые данные
            current_data = [item for item in current_data if
                            'timetable' not in item or not any(w['week_number'] == week_num for w in item['timetable'])]
            # Добавляем новые данные
            current_data.extend([item for item in pending_data if
                                 'timetable' in item and any(w['week_number'] == week_num for w in item['timetable'])])

    # Объединяем оставшиеся данные
    final_data = current_data + [item for item in pending_data if item not in current_data]

    try:
        # Сохраняем результат
        save_merged_data(final_data)

        # Очищаем временные данные
        remove_temp_data(temp_id)
        session.pop('temp_id', None)

        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@bp.errorhandler(Exception)
def handle_error(e):
    temp_id = session.get('temp_id')
    if temp_id:
        remove_temp_data(temp_id)
        session.pop('temp_id', None)
    return str(e), 500


@bp.route('/teacher/<teacher_name>/workload')
@notify_view
def teacher_workload(teacher_name):
    """Страница статистики загруженности преподавателя"""
    timetable_data = timetable_handler.read_timetable()
    time_slots = [
        {'start': '08:00', 'end': '09:20'},
        {'start': '09:30', 'end': '10:50'},
        {'start': '11:00', 'end': '12:20'},
        {'start': '12:40', 'end': '14:00'},
        {'start': '14:10', 'end': '15:30'},
        {'start': '15:40', 'end': '17:00'},
        {'start': '17:10', 'end': '18:30'},
        {'start': '18:40', 'end': '20:00'}
    ]
    workload_by_subject = {}  # Статистика по предметам
    total_stats = {  # Общая статистика
        'lectures': 0, 'practices': 0, 'labs': 0}
    group_details = {}  # Детали по группам

    # Создаем словарь для хранения информации о датах недель
    week_dates = {}

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                # Сначала собираем информацию о датах для каждой недели
                for week in data['timetable']:
                    week_number = week.get('week_number')
                    date_start = week.get('date_start')
                    date_end = week.get('date_end')
                    if week_number and date_start and date_end:
                        week_dates[week_number] = f"{date_start} - {date_end}"

                # Сортируем недели по номеру
                sorted_weeks = sorted(data['timetable'], key=lambda w: w.get('week_number', 0))

                for week in sorted_weeks:
                    week_number = week.get('week_number')

                    for group in week.get('groups', []):
                        group_name = group.get('group_name')

                        # Создаем структуры для групп если их еще нет
                        if group_name not in group_details:
                            group_details[group_name] = {
                                'total_hours': 0,
                                'lectures': 0,
                                'practices': 0,
                                'labs': 0,
                                'schedule': [],
                                'subjects': set()
                            }

                        for day in group.get('days', []):
                            weekday = day.get('weekday')
                            weekday_name = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота'][
                                weekday - 1]

                            for lesson in day.get('lessons', []):
                                if any(teacher.get('teacher_name') == teacher_name for teacher in
                                       lesson.get('teachers', [])):
                                    subject = lesson.get('subject')
                                    time_index = lesson.get('time', 0)

                                    # Получаем время начала и конца пары
                                    time_info = time_slots[time_index - 1] if 0 < time_index <= len(
                                        time_slots) else None
                                    time_str = f"{time_info['start']} - {time_info['end']}" if time_info else f"Пара {time_index}"

                                    # Добавляем даты недели к информации о занятии
                                    lesson_info = {
                                        'week': week_number,
                                        'week_dates': week_dates.get(week_number, ''),  # Добавляем даты
                                        'weekday': weekday_name,
                                        'weekday_number': weekday,
                                        'time_slot': time_index,
                                        'time_str': time_str,
                                        'subject': subject,
                                        'type': lesson.get('type'),
                                        'subgroup': lesson.get('subgroup', 0),
                                        'auditory': lesson.get('auditories', [{}])[0].get('auditory_name', '')
                                    }

                                    # Обновляем расписание группы
                                    group_details[group_name]['schedule'].append(lesson_info)
                                    group_details[group_name]['subjects'].add(subject)

                                    # Обновляем счётчики для группы
                                    group_details[group_name]['total_hours'] += 2
                                    if lesson.get('type') == 'л.':
                                        group_details[group_name]['lectures'] += 1
                                    elif lesson.get('type') == 'пр.':
                                        group_details[group_name]['practices'] += 1
                                    elif lesson.get('type') == 'лаб.':
                                        group_details[group_name]['labs'] += 1

                                    # Обновляем статистику по предметам
                                    if subject not in workload_by_subject:
                                        workload_by_subject[subject] = {
                                            'groups': set(),
                                            'lectures': 0,
                                            'practices': 0,
                                            'labs': 0,
                                            'total_hours': 0
                                        }

                                    workload_by_subject[subject]['groups'].add(group_name)
                                    workload_by_subject[subject]['total_hours'] += 2

                                    if lesson.get('type') == 'л.':
                                        workload_by_subject[subject]['lectures'] += 1
                                        total_stats['lectures'] += 1
                                    elif lesson.get('type') == 'пр.':
                                        workload_by_subject[subject]['practices'] += 1
                                        total_stats['practices'] += 1
                                    elif lesson.get('type') == 'лаб.':
                                        workload_by_subject[subject]['labs'] += 1
                                        total_stats['labs'] += 1

    # Сортируем расписание для каждой группы
    for group_data in group_details.values():
        group_data['schedule'].sort(key=lambda x: (x['week'], x['weekday_number'], x['time_slot']))
        group_data['subjects'] = sorted(list(group_data['subjects']))

    # Конвертируем множества групп в списки и сортируем
    for subject_data in workload_by_subject.values():
        subject_data['groups'] = sorted(list(subject_data['groups']))

    # Считаем общее количество часов
    total_hours = sum(subject_data['total_hours'] for subject_data in workload_by_subject.values())

    # Сортируем предметы по алфавиту
    workload_by_subject = dict(sorted(workload_by_subject.items()))

    return render_template('timetable/teacher_workload.html',
                           teacher_name=teacher_name,
                           workload_by_subject=workload_by_subject,
                           total_stats=total_stats,
                           total_hours=total_hours,
                           group_details=group_details,
                           time_slots=time_slots)


@bp.route('/room/<room_name>')
@notify_view
def room_timetable(room_name):
    timetable_data = timetable_handler.read_timetable()
    selected_week = request.args.get('week', None)

    weeks = []
    if isinstance(timetable_data, list):
        for timetable_item in timetable_data:
            if 'timetable' in timetable_item:
                for week_data in timetable_item['timetable']:
                    weeks.append(
                        {'week_number': week_data.get('week_number'), 'date_start': week_data.get('date_start'),
                         'date_end': week_data.get('date_end')})

    weeks.sort(key=lambda x: x['week_number'])

    if not selected_week and weeks:
        current_date = datetime.now()
        current_week_num = current_date.isocalendar()[1]

        closest_week = weeks[0]
        min_diff = float('inf')

        for week in weeks:
            try:
                try:
                    week_start = datetime.strptime(week['date_start'], '%d.%m.%Y')
                except ValueError:
                    week_start = datetime.strptime(week['date_start'], '%d-%m-%Y')

                week_diff = abs(week_start.isocalendar()[1] - current_week_num)
                if week_diff < min_diff:
                    min_diff = week_diff
                    closest_week = week
            except ValueError:
                continue

        selected_week = str(closest_week['week_number'])
        return redirect(url_for('timetable.room_timetable', room_name=room_name, week=selected_week))

    time_slots = [{'start': '08:00', 'end': '09:20'}, {'start': '09:30', 'end': '10:50'},
                  {'start': '11:00', 'end': '12:20'}, {'start': '12:40', 'end': '14:00'},
                  {'start': '14:10', 'end': '15:30'}, {'start': '15:40', 'end': '17:00'},
                  {'start': '17:10', 'end': '18:30'}, {'start': '18:40', 'end': '20:00'}]

    day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

    current_datetime = datetime.now()
    current_time = current_datetime.strftime('%H:%M')
    current_weekday = current_datetime.weekday() + 1

    current_pair = None
    for i, time_slot in enumerate(time_slots, 1):
        start_time = datetime.strptime(time_slot['start'], '%H:%M').time()
        end_time = datetime.strptime(time_slot['end'], '%H:%M').time()
        if start_time <= current_datetime.time() <= end_time:
            current_pair = i
            break

    def get_room_lessons(timetable, day, time):
        """Получение занятий в аудитории для конкретного дня и времени"""
        all_lessons = []
        lessons_by_type = {}

        try:
            if isinstance(timetable, list):
                for data in timetable:
                    if 'timetable' in data:
                        for week in data['timetable']:
                            if selected_week and str(week.get('week_number')) != str(selected_week):
                                continue

                            for group in week.get('groups', []):
                                group_name = group.get('group_name')
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            if (lesson.get('time') == time and any(
                                                    auditory.get('auditory_name') == room_name for auditory in
                                                    lesson.get('auditories', []))):

                                                # Создаем ключ для группировки идентичных занятий
                                                key = (lesson.get('subject'), lesson.get('type'),
                                                       str(lesson.get('teachers', [{}])[0].get('teacher_name', '')))

                                                if key not in lessons_by_type:
                                                    lesson_copy = lesson.copy()
                                                    lesson_copy['groups'] = [group_name]
                                                    lessons_by_type[key] = lesson_copy
                                                else:
                                                    lessons_by_type[key]['groups'].append(group_name)

            result = list(lessons_by_type.values())
            for lesson in result:
                lesson['groups'].sort()

            return result if result else None

        except Exception as e:
            print(f"Ошибка в get_room_lessons: {str(e)}")
            return None

    return render_template('timetable/room.html', room_name=room_name, weeks=weeks, current_week=selected_week,
                           time_slots=time_slots, day_names=day_names, timetable=timetable_data,
                           get_lessons=get_room_lessons, current_pair=current_pair, current_weekday=current_weekday,
                           current_time=current_time)


# --- Функции ---


def find_all_overlaps(timetable_data):
    """Поиск всех существующих накладок в расписании"""
    overlaps = {'room_overlaps': [],  # Накладки по аудиториям (разные предметы в одной аудитории)
                'teacher_overlaps': []  # Накладки по преподавателям
                }

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    week_number = week.get('week_number')
                    date_start = week.get('date_start')
                    date_end = week.get('date_end')

                    # Создаем временные словари для хранения занятий
                    room_lessons = {}  # key: (day, time, room) -> value: {subject: [lessons]}
                    teacher_lessons = {}  # key: (day, time, teacher) -> value: {subject_room: [lessons]}

                    for group in week.get('groups', []):
                        group_name = group.get('group_name')
                        if not group_name:
                            continue

                        for day in group.get('days', []):
                            day_number = day.get('weekday')
                            if not day_number:
                                continue

                            for lesson in day.get('lessons', []):
                                time = lesson.get('time')
                                if not time:
                                    continue

                                subject = lesson.get('subject')
                                if not subject:
                                    continue

                                subgroup = lesson.get('subgroup', 0)

                                # Проверка аудиторий
                                auditories = lesson.get('auditories', [])
                                if auditories and isinstance(auditories, list):
                                    for auditory in auditories:
                                        room_name = auditory.get('auditory_name')
                                        if room_name:
                                            room_key = (day_number, time, room_name)
                                            if room_key not in room_lessons:
                                                room_lessons[room_key] = {}

                                            if subject not in room_lessons[room_key]:
                                                room_lessons[room_key][subject] = []

                                            teachers = lesson.get('teachers', [])
                                            teacher_name = teachers[0].get(
                                                'teacher_name') if teachers else 'Нет преподавателя'

                                            room_lessons[room_key][subject].append(
                                                {'group': group_name, 'subject': subject, 'teacher': teacher_name,
                                                 'subgroup': subgroup})

                                # Проверка преподавателей
                                teachers = lesson.get('teachers', [])
                                if teachers and isinstance(teachers, list):
                                    for teacher in teachers:
                                        teacher_name = teacher.get('teacher_name')
                                        if teacher_name:
                                            teacher_key = (day_number, time, teacher_name)
                                            if teacher_key not in teacher_lessons:
                                                teacher_lessons[teacher_key] = {}

                                            # Создаем ключ из предмета и аудитории
                                            room_name = auditories[0].get(
                                                'auditory_name') if auditories else 'Нет аудитории'
                                            subject_room_key = f"{subject}_{room_name}"

                                            if subject_room_key not in teacher_lessons[teacher_key]:
                                                teacher_lessons[teacher_key][subject_room_key] = []

                                            teacher_lessons[teacher_key][subject_room_key].append(
                                                {'group': group_name, 'subject': subject, 'room': room_name,
                                                 'subgroup': subgroup})

                    # Проверяем накладки по аудиториям
                    for (day, time, room), subjects_dict in room_lessons.items():
                        if len(subjects_dict) > 1:  # Разные предметы в одной аудитории
                            all_lessons = []
                            for subject_lessons in subjects_dict.values():
                                all_lessons.extend(subject_lessons)

                            overlaps['room_overlaps'].append(
                                {'week': {'number': week_number, 'date_start': date_start, 'date_end': date_end},
                                 'day': day, 'time': time, 'room': room, 'lessons': all_lessons})

                    # Проверяем накладки по преподавателям
                    for (day, time, teacher), subject_rooms_dict in teacher_lessons.items():
                        if len(subject_rooms_dict) > 1:  # Разные предметы или аудитории
                            all_lessons = []
                            for subject_room_lessons in subject_rooms_dict.values():
                                all_lessons.extend(subject_room_lessons)

                            overlaps['teacher_overlaps'].append(
                                {'week': {'number': week_number, 'date_start': date_start, 'date_end': date_end},
                                 'day': day, 'time': time, 'teacher': teacher, 'lessons': all_lessons})

    # Сортируем накладки
    for overlap_type in ['room_overlaps', 'teacher_overlaps']:
        overlaps[overlap_type].sort(key=lambda x: (x['week']['number'], x['day'], x['time']))

    return overlaps


def export_teacher_workload(teacher_name, workload_data):
    """
    Экспортирует статистику загруженности преподавателя в Excel файл.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Загруженность преподавателя"

    # Фиксируем количество недель равным 18
    max_weeks = 18

    # Заголовок документа
    header = f"Загруженность преподавателя: {teacher_name}"
    ws.merge_cells(f'A1:{get_column_letter(max_weeks + 2)}1')  # +2 для колонки ИТОГО
    ws['A1'] = header
    ws['A1'].font = Font(name='Times New Roman', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Отступ после заголовка
    ws.append([])

    for subject, subject_data in workload_data.items():
        # Заголовок предмета
        subject_header = f"Предмет: {subject}"
        ws.append([subject_header])
        ws[f'A{ws.max_row}'].font = Font(name='Times New Roman', size=12, bold=True)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal='left', vertical='center')

        # Отступ после заголовка предмета
        ws.append([])

        for group, group_data in subject_data['groups'].items():
            # Собираем даты из расписания для каждой недели
            weeks_dates = {}
            for lesson in group_data['schedule']:
                if 0 < lesson['week'] <= max_weeks:
                    date_str = lesson.get('week_dates', '')
                    if date_str:
                        # Преобразуем формат даты
                        try:
                            dates = date_str.split(' - ')
                            start_date = datetime.strptime(dates[0], '%d-%m-%Y').strftime('%d.%m.%Y')
                            end_date = datetime.strptime(dates[1], '%d-%m-%Y').strftime('%d.%m.%Y')
                            weeks_dates[lesson['week']] = f"({start_date} {end_date})"
                        except:
                            weeks_dates[lesson['week']] = date_str

            # Заголовок группы и строка с номерами недель
            row_data = [f"Группа: {group}"]
            for week_num in range(1, max_weeks + 1):
                row_data.append(f"Неделя {week_num}")
            row_data.append("ИТОГО")  # Добавляем колонку ИТОГО
            ws.append(row_data)

            # Строка с датами
            date_row = ['']  # Пустая ячейка под названием группы
            for week_num in range(1, max_weeks + 1):
                date_row.append(weeks_dates.get(week_num, ''))
            date_row.append('')  # Пустая ячейка под ИТОГО
            ws.append(date_row)

            # Форматируем заголовки
            for col in range(1, max_weeks + 3):  # +3 для учета ИТОГО
                # Форматируем ячейку с номером недели
                week_cell = ws.cell(row=ws.max_row - 1, column=col)
                week_cell.alignment = Alignment(horizontal='center', vertical='bottom')
                week_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                          top=Side(style='thin'), bottom=Side(style='none'))
                week_cell.font = Font(name='Times New Roman', size=11, bold=True)

                # Форматируем ячейку с датой
                date_cell = ws.cell(row=ws.max_row, column=col)
                date_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                date_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                          top=Side(style='none'), bottom=Side(style='thin'))
                date_cell.font = Font(name='Times New Roman', size=9, italic=True)

            # Устанавливаем высоту строк
            ws.row_dimensions[ws.max_row - 1].height = 20
            ws.row_dimensions[ws.max_row].height = 25  # Увеличиваем для переноса текста

            # Словарь для подсчета часов по типам занятий
            hours_by_type = {
                'л.': [0] * max_weeks,
                'пр.': [0] * max_weeks,
                'лаб.': [0] * max_weeks
            }

            # Считаем часы по неделям и типам занятий
            for lesson in group_data['schedule']:
                week = lesson['week'] - 1
                lesson_type = lesson.get('type')
                if 0 <= week < max_weeks and lesson_type in hours_by_type:
                    hours_by_type[lesson_type][week] += 2

            # Заполняем данные для каждого типа занятий
            for lesson_name, lesson_code in [("Лекции", 'л.'), ("Практики", 'пр.'), ("Лабораторные", 'лаб.')]:
                row_total = 0  # Для подсчета суммы по строке
                row_data = [lesson_name] + [''] * max_weeks
                for week in range(max_weeks):
                    hours = hours_by_type[lesson_code][week]
                    if hours > 0:
                        row_data[week + 1] = hours
                        row_total += hours

                row_data.append(row_total)  # Добавляем сумму
                ws.append(row_data)

                # Форматирование строки
                row = ws.max_row
                for col in range(1, max_weeks + 3):  # +3 для учета ИТОГО
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(name='Times New Roman', size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))

                    # Выделяем колонку ИТОГО жирным
                    if col == max_weeks + 2:
                        cell.font = Font(name='Times New Roman', size=11, bold=True)

            # Отступ после таблицы группы
            ws.append([])
            ws.append([])

    # Настройка печати
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = False  # Отключаем автоматическое масштабирование
    ws.page_setup.scale = 60  # Устанавливаем масштаб в 60%
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.85

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20
    for i in range(2, max_weeks + 2):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(max_weeks + 2)].width = 12  # Ширина для ИТОГО

    # Замораживаем первую колонку
    ws.freeze_panes = 'B3'

    # Сохраняем во временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return send_file(
        tmp_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Загруженность_{teacher_name}.xlsx"
    )

def check_overlaps(timetable_data, week_number, day, time, new_lessons, group_name=None):
    """Проверяет накладки в расписании"""
    overlaps = {'room_overlaps': [], 'teacher_overlaps': [], 'total_count': 0}

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week in data['timetable']:
                    if str(week.get('week_number')) == str(week_number):
                        existing_lessons = []
                        for group in week.get('groups', []):
                            if group_name and group.get('group_name') == group_name:
                                continue

                            for day_data in group.get('days', []):
                                if day_data.get('weekday') == day:
                                    for lesson in day_data.get('lessons', []):
                                        if lesson.get('time') == time:
                                            existing_lessons.append(
                                                {'lesson': lesson, 'group': group.get('group_name')})

                        for new_lesson in new_lessons:
                            auditories = new_lesson.get('auditories', [])
                            new_room = auditories[0].get('auditory_name') if auditories else None

                            new_teachers = new_lesson.get('teachers', [])
                            new_teacher = new_teachers[0].get('teacher_name') if new_teachers else None
                            new_subject = new_lesson.get('subject')

                            if not new_room or not new_teacher or not new_subject:
                                continue

                            for existing in existing_lessons:
                                existing_lesson = existing['lesson']

                                existing_auditories = existing_lesson.get('auditories', [])
                                existing_room = existing_auditories[0].get(
                                    'auditory_name') if existing_auditories else None

                                existing_teachers = existing_lesson.get('teachers', [])
                                existing_teacher = existing_teachers[0].get(
                                    'teacher_name') if existing_teachers else None
                                existing_subject = existing_lesson.get('subject')

                                if not existing_room or not existing_teacher or not existing_subject:
                                    continue

                                # Проверка накладок по аудиториям (только если разные предметы)
                                if new_room == existing_room and new_subject != existing_subject:
                                    overlap = {'type': 'room', 'room': new_room, 'time': time, 'day': day, 'lessons': [
                                        {'subject': existing_subject, 'group': existing['group'],
                                         'teacher': existing_teacher},
                                        {'subject': new_subject, 'group': group_name, 'teacher': new_teacher}]}
                                    overlaps['room_overlaps'].append(overlap)
                                    overlaps['total_count'] += 1

                                # Проверка накладок по преподавателям (если разные предметы или разные аудитории)
                                if new_teacher == existing_teacher and (
                                        new_subject != existing_subject or new_room != existing_room):
                                    overlap = {'type': 'teacher', 'teacher': new_teacher, 'time': time, 'day': day,
                                               'lessons': [{'subject': existing_subject, 'group': existing['group'],
                                                            'room': existing_room},
                                                           {'subject': new_subject, 'group': group_name,
                                                            'room': new_room}]}
                                    overlaps['teacher_overlaps'].append(overlap)
                                    overlaps['total_count'] += 1

    return overlaps


def merge_weeks(weeks):
    """
    Объединяет данные о неделях расписания из разных файлов в единый согласованный формат.

    Функция обрабатывает списки недель, группируя их по номеру недели и объединяя данные о группах,
    днях и занятиях. При объединении сохраняется уникальность записей и избегается дублирование данных.

    Args:
        weeks (list): Список словарей с данными о неделях. Каждая неделя должна содержать:
            - week_number (int): Номер недели
            - date_start (str): Дата начала недели
            - date_end (str): Дата окончания недели
            - groups (list): Список групп с их расписанием

    Returns:
        list: Список объединенных недель, где каждая неделя содержит:
            - week_number: Номер недели
            - date_start: Дата начала недели
            - date_end: Дата окончания недели
            - groups: Список всех групп с объединенным расписанием

    Note:
        - Функция сохраняет все уникальные занятия для каждой группы
        - При совпадении времени и подгруппы занятия объединяются
        - Ведется подробное логирование процесса объединения
    """

    print("\n=== Начало объединения недель ===")

    merged_result = []
    weeks_by_number = {}

    # Группируем недели по номеру
    for week in weeks:
        week_num = week['week_number']
        if week_num not in weeks_by_number:
            weeks_by_number[week_num] = []
        weeks_by_number[week_num].append(week)
        print(f"Добавлена неделя {week_num} в группировку")

    # Объединяем данные для каждого номера недели
    for week_num, week_list in weeks_by_number.items():
        print(f"\nОбработка недели {week_num}")

        merged_week = {'week_number': week_num, 'date_start': week_list[0]['date_start'],
                       'date_end': week_list[0]['date_end'], 'groups': []}

        # Словарь для хранения групп
        groups_dict = {}

        # Собираем все группы из всех файлов
        for week in week_list:
            for group in week.get('groups', []):
                group_name = group['group_name']
                print(f"Обработка группы {group_name}")

                if group_name not in groups_dict:
                    groups_dict[group_name] = {'group_name': group_name, 'days': []}
                    # Инициализируем дни недели
                    for i in range(1, 7):
                        groups_dict[group_name]['days'].append({'weekday': i, 'lessons': []})
                    print(f"Создана новая запись для группы {group_name}")

                # Обрабатываем каждый день группы
                for day in group.get('days', []):
                    weekday = day.get('weekday')
                    if weekday:
                        day_index = weekday - 1
                        if day_index < len(groups_dict[group_name]['days']):
                            existing_day = groups_dict[group_name]['days'][day_index]

                            # Обрабатываем каждый урок
                            for new_lesson in day.get('lessons', []):
                                should_add = True
                                new_time = new_lesson.get('time')
                                new_subgroup = new_lesson.get('subgroup', 0)

                                # Проверяем существующие уроки
                                for existing_lesson in existing_day['lessons']:
                                    if (existing_lesson.get('time') == new_time and existing_lesson.get(
                                            'subgroup') == new_subgroup and existing_lesson.get(
                                        'subject') == new_lesson.get('subject')):
                                        should_add = False
                                        break

                                if should_add:
                                    existing_day['lessons'].append(new_lesson)
                                    print(f"Добавлен урок для группы {group_name}, день {weekday}, "
                                          f"время {new_time}, подгруппа {new_subgroup}")

        # Преобразуем словарь групп обратно в список
        merged_week['groups'] = list(groups_dict.values())
        merged_result.append(merged_week)
        print(f"Неделя {week_num} объединена")

    print("=== Завершено объединение недель ===\n")
    return merged_result


def save_temp_data(data):
    """
    Сохраняет временные данные в файл с уникальным идентификатором.

    Создает временный файл с использованием UUID для хранения данных при обработке
    конфликтов или промежуточных результатов работы системы.

    Args:
        data (any): Данные для сохранения во временный файл (будут сериализованы через pickle)

    Returns:
        str: Уникальный идентификатор (UUID) сохраненных данных

    Note:
        Файл создается в системной временной директории с префиксом 'timetable_temp_'
    """

    temp_id = str(uuid.uuid4())
    temp_path = os.path.join(tempfile.gettempdir(), f'timetable_temp_{temp_id}.pkl')

    with open(temp_path, 'wb') as f:
        pickle.dump(data, f)

    return temp_id


def find_transfer_slots(timetable_data, group_name, lesson_data, current_week, current_day, current_time):
    transfer_options = []
    current_week_num = int(current_week)

    # Получаем список доступных недель
    available_weeks = []
    if isinstance(timetable_data, list):
        for data in timetable_data:
            for week in data.get('timetable', []):
                week_num = week.get('week_number')
                if week_num >= current_week_num:
                    # Проверяем, есть ли занятия на этой неделе
                    has_lessons = False
                    for group in week.get('groups', []):
                        if group.get('days', []):
                            has_lessons = True
                            break
                    if has_lessons:
                        available_weeks.append(week_num)

    # Для каждой доступной недели
    for week_num in available_weeks:
        # Определяем диапазон дней для проверки
        start_day = current_day if week_num == current_week_num else 1

        for day in range(start_day, 7):
            for time in range(1, 9):
                # Пропускаем текущий слот
                if week_num == current_week_num and day == current_day and time == current_time:
                    continue

                score = calculate_transfer_score(timetable_data, week_num, day, time, current_week_num, current_day,
                                                 current_time, group_name, lesson_data)

                if score['total'] > 0:
                    transfer_options.append(
                        {'week': week_num, 'day': day, 'time': time, 'score': score['total'], 'room': score['room'],
                         'conflicts': score['conflicts']})

    transfer_options.sort(key=lambda x: x['score'], reverse=True)
    return transfer_options[:5]


def calculate_transfer_score(timetable_data, week_num, day, time, current_week, current_day, current_time, group_name,
                             lesson_data):
    """
    Рассчитывает рейтинг для варианта переноса
    """
    score = {'total': 0, 'conflicts': [], 'room': lesson_data.get('auditories', [{}])[0].get('auditory_name', '')}

    teacher_name = lesson_data.get('teachers', [{}])[0].get('teacher_name')
    room_name = score['room']

    # Базовые баллы
    if week_num == current_week:
        score['total'] += 20  # Текущая неделя

    if day == current_day:
        score['total'] += 15  # Тот же день
    elif abs(day - current_day) == 1:
        score['total'] += 10  # Соседний день

    if abs(time - current_time) == 1:
        score['total'] += 10  # Соседняя пара

    # Проверяем доступность
    if not is_time_free(timetable_data, week_num, day, time, group_name=group_name):
        return {'total': 0, 'conflicts': ['Группа занята'], 'room': room_name}

    if not is_time_free(timetable_data, week_num, day, time, teacher_name=teacher_name):
        score['conflicts'].append(f'Преподаватель {teacher_name} занят')
        score['total'] -= 30

    if not is_time_free(timetable_data, week_num, day, time, room_name=room_name):
        similar_rooms = find_similar_rooms(timetable_data, room_name, week_num, day, time)
        if similar_rooms:
            score['room'] = similar_rooms[0]
            score['total'] += 5
        else:
            score['conflicts'].append(f'Аудитория {room_name} занята')
            score['total'] -= 20

    # Проверяем "окна"
    windows = count_windows(timetable_data, week_num, day, group_name)
    score['total'] -= windows * 5

    return score


def is_time_free(timetable_data, week, day, time, group_name=None, teacher_name=None, room_name=None):
    """
    Проверяет доступность слота для группы/преподавателя/аудитории
    """
    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week_data in data['timetable']:
                    if str(week_data.get('week_number')) == str(week):
                        for group in week_data.get('groups', []):
                            # Проверяем для группы
                            if group_name and group.get('group_name') == group_name:
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            if lesson.get('time') == time:
                                                return False

                            # Проверяем для преподавателя
                            if teacher_name:
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            if lesson.get('time') == time:
                                                for teacher in lesson.get('teachers', []):
                                                    if teacher.get('teacher_name') == teacher_name:
                                                        return False

                            # Проверяем для аудитории
                            if room_name:
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            if lesson.get('time') == time:
                                                for room in lesson.get('auditories', []):
                                                    if room.get('auditory_name') == room_name:
                                                        return False
    return True


def find_similar_rooms(timetable_data, room_name, week, day, time):
    """
    Ищет похожие свободные аудитории
    """
    similar_rooms = []

    # Получаем корпус и номер аудитории
    try:
        building = room_name.split('.')[0]
    except:
        building = None

    for data in timetable_data:
        if 'timetable' in data:
            for week_data in data['timetable']:
                for group in week_data.get('groups', []):
                    for day_data in group.get('days', []):
                        for lesson in day_data.get('lessons', []):
                            for room in lesson.get('auditories', []):
                                candidate = room.get('auditory_name')
                                if candidate and candidate != room_name:
                                    # Проверяем тот же корпус
                                    try:
                                        if building and candidate.split('.')[0] == building:
                                            if is_time_free(timetable_data, week, day, time, room_name=candidate):
                                                similar_rooms.append(candidate)
                                    except:
                                        continue

    return list(set(similar_rooms))


def count_windows(timetable_data, week, day, group_name):
    """
    Подсчитывает количество "окон" в расписании группы
    """
    lessons_times = set()

    if isinstance(timetable_data, list):
        for data in timetable_data:
            if 'timetable' in data:
                for week_data in data['timetable']:
                    if str(week_data.get('week_number')) == str(week):
                        for group in week_data.get('groups', []):
                            if group.get('group_name') == group_name:
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        for lesson in day_data.get('lessons', []):
                                            lessons_times.add(lesson.get('time'))

    if lessons_times:
        times = sorted(list(lessons_times))
        windows = 0
        for i in range(len(times) - 1):
            if times[i + 1] - times[i] > 1:
                windows += times[i + 1] - times[i] - 1
        return windows
    return 0


def load_temp_data(temp_id):
    """
    Загружает данные из временного файла по его идентификатору.

    Args:
        temp_id (str): Уникальный идентификатор временного файла

    Returns:
        any: Десериализованные данные из временного файла или None, если файл не найден

    Note:
        Осуществляет поиск файла в системной временной директории
        Возвращает None, если файл не существует или возникла ошибка при чтении
    """

    temp_path = os.path.join(tempfile.gettempdir(), f'timetable_temp_{temp_id}.pkl')

    if os.path.exists(temp_path):
        with open(temp_path, 'rb') as f:
            data = pickle.load(f)
        return data
    return None


def remove_temp_data(temp_id):
    """
    Удаляет временный файл по его идентификатору.

    Очищает временные данные после их использования для освобождения системных ресурсов.

    Args:
        temp_id (str): Уникальный идентификатор временного файла

    Note:
        Проверяет существование файла перед удалением
        Безопасно игнорирует отсутствие файла
    """

    temp_path = os.path.join(tempfile.gettempdir(), f'timetable_temp_{temp_id}.pkl')
    if os.path.exists(temp_path):
        os.remove(temp_path)


class ExcelExporter:
    def __init__(self):
        # Расписание звонков
        self.time_slots = ['08:00 - 09:20', '09:30 - 10:50', '11:00 - 12:20', '12:40 - 14:00', '14:10 - 15:30',
                           '15:40 - 17:00', '17:10 - 18:30', '18:40 - 20:00']
        self.days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']

        # Определение стилей
        self.styles = {'header': {'fill': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
                                  'font': Font(bold=True),
                                  'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                                                   top=Side(style='thin'), bottom=Side(style='thin')),
                                  'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)},
                       'cell': {
                           'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                            bottom=Side(style='thin')),  # Обновляем выравнивание для лучшей читаемости
                           'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True,
                                                  shrink_to_fit=False  # Отключаем уменьшение текста
                                                  )},
                       'title': {'font': Font(bold=True, size=14), 'alignment': Alignment(horizontal='center')},
                       'lesson_types': {'л.': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),
                                        'пр.': PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
                                        'лаб.': PatternFill(start_color='FFE6FF', end_color='FFE6FF',
                                                            fill_type='solid')}}

    def create_excel(self, timetable_data, week_number, group_name=None, teacher_name=None, room_name=None):
        """Создание Excel файла с расписанием"""
        wb = Workbook()
        ws = wb.active

        # Находим информацию о выбранной неделе
        week_info = self._get_week_info(timetable_data, week_number)

        # Настраиваем заголовок и общую информацию
        self._setup_worksheet_info(ws, week_info, group_name, teacher_name, room_name)

        # Создаем шапку таблицы
        self._create_table_header(ws, week_info)

        # Заполняем временные слоты
        self._fill_time_slots(ws)

        # Настраиваем размеры столбцов
        self._setup_column_dimensions(ws)

        # Заполняем данные расписания
        if week_info:
            self._fill_timetable_data(ws, week_info, group_name, teacher_name, room_name)

        # Применяем финальное форматирование
        self._apply_final_formatting(ws)

        self._apply_cell_formatting(ws)

        return wb

    def _get_week_info(self, timetable_data, week_number):
        """Получение информации о выбранной неделе"""
        if isinstance(timetable_data, list) and len(timetable_data) > 0:
            for week in timetable_data[0].get('timetable', []):
                if str(week.get('week_number')) == str(week_number):
                    return week
        return None

    def _setup_worksheet_info(self, ws, week_info, group_name, teacher_name, room_name):
        """Настройка общей информации листа"""
        # Устанавливаем название листа
        if group_name:
            ws.title = f"Группа {group_name}"
        elif teacher_name:
            ws.title = f"Преподаватель {teacher_name}"
        elif room_name:
            ws.title = f"Аудитория {room_name}"

        # Формируем заголовок
        title_parts = []
        if group_name:
            title_parts.append(f'Расписание группы {group_name}')
        elif teacher_name:
            title_parts.append(f'Расписание преподавателя {teacher_name}')
        elif room_name:
            title_parts.append(f'Расписание аудитории {room_name}')

        if week_info:
            title_parts.append(
                f'Неделя {week_info.get("week_number")} ({week_info.get("date_start")} - {week_info.get("date_end")})')

        # Применяем заголовок
        ws.merge_cells('A1:G1')
        ws['A1'] = ' | '.join(title_parts)
        ws['A1'].font = self.styles['title']['font']
        ws['A1'].alignment = self.styles['title']['alignment']

    def _create_table_header(self, ws, week_info):
        """Создание шапки таблицы"""
        ws['A2'] = 'Время'
        if week_info and week_info.get('date_start'):
            try:
                start = datetime.strptime(week_info['date_start'], '%d.%m.%Y')
                for i, day in enumerate(self.days):
                    col = get_column_letter(i + 2)
                    current_date = start + timedelta(days=i)
                    ws[f'{col}2'] = f'{day}\n{current_date.strftime("%d.%m")}'
            except:
                self._fill_default_headers(ws)
        else:
            self._fill_default_headers(ws)

        # Применяем стили к заголовкам
        for i in range(1, 8):
            cell = ws[f'{get_column_letter(i)}2']
            cell.fill = self.styles['header']['fill']
            cell.border = self.styles['header']['border']
            cell.alignment = self.styles['header']['alignment']
            cell.font = self.styles['header']['font']

    def _fill_default_headers(self, ws):
        """Заполнение заголовков по умолчанию"""
        for i, day in enumerate(self.days):
            ws[f'{get_column_letter(i + 2)}2'] = day

    def _fill_time_slots(self, ws):
        """Заполнение временных слотов"""
        for i, time_slot in enumerate(self.time_slots, start=3):
            cell = ws[f'A{i}']
            cell.value = time_slot
            cell.border = self.styles['cell']['border']
            cell.alignment = self.styles['cell']['alignment']

    def _setup_column_dimensions(self, ws):
        """Настройка размеров столбцов"""
        ws.column_dimensions['A'].width = 15
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 30

    def _fill_timetable_data(self, ws, week_info, group_name, teacher_name, room_name):
        """Заполнение данных расписания"""
        if group_name:
            self._fill_group_data(ws, week_info, group_name)
        elif teacher_name:
            self._fill_teacher_data(ws, week_info, teacher_name)
        elif room_name:
            self._fill_room_data(ws, week_info, room_name)

    def _fill_group_data(self, ws, week, group_name):
        """Заполнение данных для группы"""
        for group in week.get('groups', []):
            if group.get('group_name') == group_name:
                for day in group.get('days', []):
                    day_index = day.get('weekday', 0)
                    # Сначала группируем занятия по времени
                    lessons_by_time = {}
                    for lesson in day.get('lessons', []):
                        time_index = lesson.get('time', 0)
                        if time_index not in lessons_by_time:
                            lessons_by_time[time_index] = []
                        lessons_by_time[time_index].append(lesson)

                    # Теперь обрабатываем сгруппированные занятия
                    for time_index, lessons in lessons_by_time.items():
                        if 0 < time_index <= 8 and 0 < day_index <= 6:
                            cell = ws.cell(row=time_index + 2, column=day_index + 1)

                            if len(lessons) == 1:
                                # Одно занятие
                                cell.value = self._format_lesson(lessons[0])
                                self._apply_lesson_style(cell, lessons[0])
                            else:
                                # Несколько занятий (подгруппы)
                                # Сортируем по номеру подгруппы
                                lessons.sort(key=lambda x: x.get('subgroup', 0))
                                formatted_lessons = []
                                for lesson in lessons:
                                    formatted_lessons.append(self._format_lesson(lesson))
                                cell.value = '\n---\n'.join(formatted_lessons)
                                # Применяем смешанный стиль или стиль первого занятия
                                self._apply_multiple_lessons_style(cell, lessons)

    def _fill_teacher_data(self, ws, week, teacher_name):
        """Заполнение данных для преподавателя"""
        for group in week.get('groups', []):
            for day in group.get('days', []):
                day_index = day.get('weekday', 0)
                lessons_by_time = {}
                for lesson in day.get('lessons', []):
                    if any(teacher.get('teacher_name') == teacher_name for teacher in lesson.get('teachers', [])):
                        time_index = lesson.get('time', 0)
                        if time_index not in lessons_by_time:
                            lessons_by_time[time_index] = []
                        lessons_by_time[time_index].append((lesson, group.get('group_name')))

                for time_index, lesson_group_pairs in lessons_by_time.items():
                    if 0 < time_index <= 8 and 0 < day_index <= 6:
                        cell = ws.cell(row=time_index + 2, column=day_index + 1)
                        formatted_lessons = []
                        for lesson, group_name in lesson_group_pairs:
                            lesson_text = f"{lesson.get('subject')}\n{group_name}"
                            if lesson.get('subgroup'):
                                lesson_text += f"\nПодгруппа {lesson.get('subgroup')}"
                            lesson_text += f"\nауд. {lesson.get('auditories', [{}])[0].get('auditory_name', '')}"
                            formatted_lessons.append(lesson_text)
                        cell.value = '\n---\n'.join(formatted_lessons)
                        self._apply_multiple_lessons_style(cell, [pair[0] for pair in lesson_group_pairs])

    def _fill_room_data(self, ws, week, room_name):
        """Заполнение данных для аудитории"""
        for group in week.get('groups', []):
            for day in group.get('days', []):
                day_index = day.get('weekday', 0)
                lessons_by_time = {}
                for lesson in day.get('lessons', []):
                    if any(room.get('auditory_name') == room_name for room in lesson.get('auditories', [])):
                        time_index = lesson.get('time', 0)
                        if time_index not in lessons_by_time:
                            lessons_by_time[time_index] = []
                        lessons_by_time[time_index].append((lesson, group.get('group_name')))

                for time_index, lesson_group_pairs in lessons_by_time.items():
                    if 0 < time_index <= 8 and 0 < day_index <= 6:
                        cell = ws.cell(row=time_index + 2, column=day_index + 1)
                        formatted_lessons = []
                        for lesson, group_name in lesson_group_pairs:
                            lesson_text = f"{lesson.get('subject')}\n{group_name}"
                            if lesson.get('subgroup'):
                                lesson_text += f"\nПодгруппа {lesson.get('subgroup')}"
                            lesson_text += f"\n{lesson.get('teachers', [{}])[0].get('teacher_name', '')}"
                            formatted_lessons.append(lesson_text)
                        cell.value = '\n---\n'.join(formatted_lessons)
                        self._apply_multiple_lessons_style(cell, [pair[0] for pair in lesson_group_pairs])

    def _format_lesson(self, lesson):
        """Форматирование информации о занятии в компактном виде"""
        lines = []

        # Первая строка: предмет
        subject_line = lesson.get('subject', '')
        lines.append(subject_line)

        # Вторая строка: тип занятия и подгруппа
        type_subgroup_parts = []
        if lesson.get('type'):
            type_subgroup_parts.append(lesson.get('type'))
        if lesson.get('subgroup'):
            type_subgroup_parts.append(f"Подгруппа {lesson.get('subgroup')}")
        if type_subgroup_parts:
            lines.append(' • '.join(type_subgroup_parts))

        # Третья строка: преподаватель
        if lesson.get('teachers'):
            teacher_name = lesson.get('teachers', [{}])[0].get('teacher_name', '')
            if teacher_name:
                lines.append(teacher_name)

        # Четвертая строка: аудитория
        if lesson.get('auditories'):
            auditory_name = lesson.get('auditories', [{}])[0].get('auditory_name', '')
            if auditory_name:
                lines.append(f"ауд. {auditory_name}")

        return '\n'.join(lines)

    def _apply_lesson_style(self, cell, lesson):
        """Применение стилей к ячейке с занятием"""
        cell.border = self.styles['cell']['border']

        # Обновленное выравнивание для компактности
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)

        # Применяем цвет в зависимости от типа занятия
        lesson_type = lesson.get('type')
        if lesson_type in self.styles['lesson_types']:
            cell.fill = self.styles['lesson_types'][lesson_type]

        # Устанавливаем шрифт
        cell.font = Font(size=9)  # Немного уменьшаем размер шрифта для компактности

    def _apply_cell_formatting(self, ws):
        """Применение форматирования к ячейкам"""
        # Настройка размеров
        ws.column_dimensions['A'].width = 15
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 35  # Немного уменьшаем ширину

        # Настройка высоты строк
        for row in range(3, 11):  # Строки с занятиями
            max_lines = 1
            for col in range(2, 8):  # Все дни недели
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    lines_count = len(str(cell.value).split('\n'))
                    max_lines = max(max_lines, lines_count)

            # Устанавливаем высоту строки на основе количества строк текста
            row_height = max(15 * max_lines, 30)  # Минимум 30, иначе 15 пикселей на строку
            ws.row_dimensions[row].height = row_height

    def _apply_multiple_lessons_style(self, cell, lessons):
        """Применение стилей к ячейке с несколькими занятиями"""
        cell.border = self.styles['cell']['border']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)

        # Устанавливаем шрифт
        cell.font = Font(size=9)

        # Если несколько занятий, добавляем разделитель
        if len(lessons) > 1:
            cell_text = cell.value
            if cell_text:
                parts = cell_text.split('\n---\n')
                cell.value = '\n--\n'.join(parts)  # Используем более компактный разделитель

    def _apply_final_formatting(self, ws):
        """Применение финального форматирования к таблице"""
        for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=7):
            for cell in row:
                if not cell.fill.start_color.index:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                if not cell.border:
                    cell.border = self.styles['cell']['border']
                if not cell.alignment:
                    cell.alignment = self.styles['cell']['alignment']

        # Устанавливаем высоту строк и ширину столбцов
        ws.column_dimensions['A'].width = 15
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 40  # Увеличиваем ширину столбцов

        # Устанавливаем автоматическую высоту строк
        for row in range(2, 11):
            ws.row_dimensions[row].height = None  # Автоматическая высота
            # Устанавливаем минимальную высоту, если содержимое маленькое
            current_height = ws.row_dimensions[row].height
            if current_height is None or current_height < 45:
                ws.row_dimensions[row].height = 45


def get_unique_values(timetable_data):
    """
    Извлекает уникальные значения для различных параметров расписания.

    Обрабатывает данные расписания и собирает уникальные значения для предметов,
    преподавателей, аудиторий и типов занятий.

    Args:
        timetable_data (list): Список данных расписания

    Returns:
        dict: Словарь с уникальными значениями:
            - subjects (list): Список уникальных предметов
            - teachers (list): Список уникальных преподавателей
            - auditories (list): Список уникальных аудиторий
            - lesson_types (list): Список типов занятий

    Note:
        Все списки сортируются в алфавитном порядке
        Пустые значения фильтруются
    """
    subjects = set()
    teachers = set()
    auditories = set()

    if isinstance(timetable_data, list) and len(timetable_data) > 0:
        weeks_data = timetable_data[0].get('timetable', [])
        for week in weeks_data:
            for group in week.get('groups', []):
                for day in group.get('days', []):
                    for lesson in day.get('lessons', []):
                        subjects.add(lesson.get('subject'))
                        for teacher in lesson.get('teachers', []):
                            teachers.add(teacher.get('teacher_name'))
                        for auditory in lesson.get('auditories', []):
                            auditories.add(auditory.get('auditory_name'))

    return {'subjects': sorted(list(filter(None, subjects))), 'teachers': sorted(list(filter(None, teachers))),
            'auditories': sorted(list(filter(None, auditories))), 'lesson_types': ['л.', 'пр.', 'лаб.']}


def get_group_timetable(group_name, timetable_data, selected_week=None):
    """
    Получает расписание для конкретной группы на выбранную неделю.

    Args:
        group_name (str): Название группы
        timetable_data (list): Данные расписания
        selected_week (str, optional): Номер выбранной недели. По умолчанию None

    Returns:
        tuple: (week_info, group_info)
            - week_info: Информация о неделе или None
            - group_info: Информация о группе или None

    Note:
        Если неделя не указана, берется первая доступная неделя
        Возвращает None, None если группа или неделя не найдены
    """

    if isinstance(timetable_data, list) and len(timetable_data) > 0:
        # Получаем данные из первого элемента списка
        weeks_data = timetable_data[0].get('timetable', [])

        # Если неделя не выбрана, берем первую
        if selected_week is None and weeks_data:
            selected_week = str(weeks_data[0].get('week_number'))

        # Ищем данные для выбранной недели
        for week in weeks_data:
            if str(week.get('week_number')) == str(selected_week):
                # Ищем конкретную группу
                for group in week.get('groups', []):
                    if group.get('group_name') == group_name:
                        return week, group
    return None, None


def get_lessons(timetable, day, time, group_name, selected_week=None):
    """
    Извлекает информацию о занятиях для конкретной группы, дня и времени.

    Args:
        timetable (list): Данные расписания
        day (int): Номер дня недели (1-6)
        time (int): Номер пары (1-8)
        group_name (str): Название группы
        selected_week (str, optional): Номер недели. По умолчанию None

    Returns:
        list или None: Список занятий, отсортированный по номеру подгруппы, или None если занятий нет

    Note:
        Занятия сортируются по номеру подгруппы (общие занятия идут первыми)
        Обрабатывает ошибки и логирует их
    """

    try:
        if isinstance(timetable, list):
            for data in timetable:
                if 'timetable' in data:
                    for week in data['timetable']:
                        # Проверяем номер недели, если он указан
                        if selected_week and str(week.get('week_number')) != str(selected_week):
                            continue

                        for group in week.get('groups', []):
                            if group.get('group_name') == group_name:
                                for day_data in group.get('days', []):
                                    if day_data.get('weekday') == day:
                                        # Собираем все занятия для данного времени
                                        all_lessons = [lesson for lesson in day_data.get('lessons', []) if
                                                       lesson.get('time') == time]

                                        # Если есть занятия, сортируем их по номеру подгруппы
                                        if all_lessons:
                                            # Сортировка по номеру подгруппы (0 - общие занятия идут первыми)
                                            all_lessons.sort(key=lambda x: x.get('subgroup', 0))
                                            return all_lessons
        return None
    except Exception as e:
        print(f"Ошибка в get_lessons: {str(e)}")
        return None


def update_lessons(self, group_name, day, time, new_lessons):
    """
    Обновление занятий в расписании с учетом номера недели
    """
    try:
        print("\n=== Начало обновления занятий ===")
        print(f"Параметры: группа={group_name}, день={day}, время={time}")
        print(f"Новые занятия: {json.dumps(new_lessons, ensure_ascii=False, indent=2)}")

        # Читаем данные
        data = self.read_timetable()
        if not data or not isinstance(data, list):
            print("Ошибка: неверный формат данных")
            return False

        # Создаем резервную копию
        self._create_backup()

        # Получаем номер недели из URL или параметров запроса
        from flask import request
        week_number = request.args.get('week')
        if not week_number:
            print("Ошибка: не указан номер недели")
            return False

        week_number = int(week_number)
        print(f"Номер недели: {week_number}")

        # Обновляем данные с учетом недели
        found = False
        timetable_data = data[0].get('timetable', [])

        for week in timetable_data:
            if week.get('week_number') != week_number:
                continue

            print(f"Обрабатываем неделю {week_number}")
            for group in week.get('groups', []):
                if group.get('group_name') != group_name:
                    continue

                print(f"Найдена группа {group_name}")
                for day_data in group.get('days', []):
                    if day_data.get('weekday') != day:
                        continue

                    print(f"Найден день {day}")
                    # Сохраняем существующие занятия для других недель
                    existing_lessons = day_data.get('lessons', [])
                    other_weeks_lessons = [lesson for lesson in existing_lessons if
                                           lesson.get('time') == time and lesson.get('week',
                                                                                     week_number) != week_number]

                    # Обновляем занятия для текущей недели
                    current_week_lessons = [lesson for lesson in existing_lessons if
                                            lesson.get('time') != time or lesson.get('week',
                                                                                     week_number) != week_number]

                    # Добавляем новые занятия с указанием недели
                    if new_lessons:
                        for lesson in new_lessons:
                            lesson_copy = lesson.copy()
                            lesson_copy['week'] = week_number
                            current_week_lessons.append(lesson_copy)

                    # Объединяем все занятия
                    day_data['lessons'] = current_week_lessons + other_weeks_lessons

                    # Сортируем занятия
                    day_data['lessons'].sort(key=lambda x: (x.get('week', 0), x.get('time', 0), x.get('subgroup', 0)))

                    found = True
                    print(f"Обновлено занятий: {len(new_lessons) if new_lessons else 0}")
                    break
                if found:
                    break
            if found:
                break

        if not found:
            print("Ошибка: занятия не найдены")
            return False

        # Сохраняем обновленные данные
        success = self.save_timetable(data)
        print(f"Результат сохранения: {success}")
        return success

    except Exception as e:
        print(f"Ошибка в update_lessons: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False


def read_merged_file():
    """
    Читает объединенный файл расписания с поддержкой различных кодировок.

    Пытается прочитать JSON файл в различных кодировках и обрабатывает возможные ошибки.

    Returns:
        list: Список данных расписания или пустой список в случае ошибки

    Supported encodings:
        - windows-1251
        - utf-8
        - utf-8-sig

    Note:
        Обрабатывает ошибки декодирования и парсинга JSON
        Логирует ошибки для отладки
    """
    try:
        if os.path.exists(MERGED_FILE):
            encodings = ['windows-1251', 'utf-8', 'utf-8-sig']

            for encoding in encodings:
                try:
                    with open(MERGED_FILE, 'r', encoding=encoding) as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            return data
                        return [data]
                except UnicodeDecodeError:
                    continue
                except json.JSONDecodeError as e:
                    print(f"Ошибка парсинга JSON с {encoding}: {str(e)}")
                    continue

    except Exception as e:
        print(f"Ошибка чтения файла: {str(e)}")
        return []

    return []


def save_merged_data(data):
    """
    Сохраняет объединенные данные расписания в файл.

    Создает необходимые директории, объединяет недели и сохраняет данные
    в файл с определенной кодировкой.

    Args:
        data (list): Список данных для сохранения

    Returns:
        bool: True если сохранение успешно, False в случае ошибки

    Note:
        Создает директории при необходимости
        Объединяет недели перед сохранением
        Использует кодировку windows-1251
        Ведет подробное логирование процесса
    """

    print("\n=== Начало сохранения данных ===")
    try:
        # Создаем папку, если её нет
        os.makedirs(os.path.dirname(MERGED_FILE), exist_ok=True)
        print(f"Директория проверена: {os.path.dirname(MERGED_FILE)}")

        # Объединяем недели
        weeks = []
        for item in data:
            if 'timetable' in item:
                weeks.extend(item['timetable'])
        print(f"Собрано недель для объединения: {len(weeks)}")

        # Объединяем недели
        merged_weeks = merge_weeks(weeks)
        print(f"Объединено недель: {len(merged_weeks)}")

        # Сохраняем результат
        final_data = [{'timetable': merged_weeks}]

        with open(MERGED_FILE, 'w', encoding='windows-1251') as f:
            json.dump(final_data, f, ensure_ascii=False, indent=2)
        print(f"Данные сохранены в файл: {MERGED_FILE}")

        return True
    except Exception as e:
        print(f"Ошибка сохранения: {str(e)}")
        return False
